"""
export_metrics.py — Read cached parquet files, compute key metrics, export to Excel

Reads every ticker's parquet files from xbrl_cache/, computes the same metrics
the pipeline computes in agents.py, and writes one Excel workbook with:

    Sheet 1: Summary        — one row per ticker, all key metrics
    Sheet 2: Gross Margin   — FY values across all periods
    Sheet 3: Operating Margin
    Sheet 4: Net Margin
    Sheet 5: ROE / ROA
    Sheet 6: D/E & Leverage
    Sheet 7: Failures       — tickers with missing/unresolvable data

Usage
-----
    python export_metrics.py
    python export_metrics.py --out my_results.xlsx
    python export_metrics.py --ticker AAPL MSFT NVDA   # subset
"""

import os
import sys
import argparse
import sqlite3
import datetime
import warnings
warnings.filterwarnings("ignore")

import pandas as pd

PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
XBRL_DIR    = os.path.join(PROJECT_DIR, "xbrl_cache")
CACHE_DB    = os.path.join(PROJECT_DIR, "filing_cache.db")
DEFAULT_OUT = os.path.join(PROJECT_DIR, f"sp500_metrics_{datetime.date.today().isoformat()}.xlsx")


# ── Metric extraction ─────────────────────────────────────────────────────────

# Map standard_concept names to readable labels
CONCEPTS = {
    "Revenue":           "Revenue",
    "GrossProfit":       "GrossProfit",
    "OperatingIncome":   "OperatingIncome",
    "NetIncome":         "NetIncome",
    "EBITDA":            "EBITDA",
    "TotalAssets":       "TotalAssets",
    "StockholdersEquity":"StockholdersEquity",
    "TotalDebt":         "TotalDebt",
    "LongTermDebt":      "LongTermDebt",
    "CurrentAssets":     "CurrentAssets",
    "CurrentLiabilities":"CurrentLiabilities",
    "CashAndEquivalents":"CashAndEquivalents",
    "OperatingCashFlow": "OperatingCashFlow",
    "Capex":             "Capex",
    "DilutedShares":     "DilutedShares",
}


def get_row_values(df: pd.DataFrame, concept: str, use_max: bool = False) -> dict:
    """
    Return {period_label: value} for a given standard_concept.
    Period labels are the date column headers (e.g. '2025-12-31 (FY)').
    use_max=True: mirrors _get_max_vector — takes the row with the largest
    absolute value per period. Use for OCF, Assets, Equity where a subtotal
    row may appear before the consolidated total.
    """
    if df is None or df.empty:
        return {}
    meta_cols = {"standard_concept", "concept", "label", "unit"}
    date_cols = [c for c in df.columns if c not in meta_cols]
    rows = df[df["standard_concept"] == concept]
    if rows.empty:
        return {}

    if not use_max or len(rows) == 1:
        row = rows.iloc[0]
        result = {}
        for col in date_cols:
            v = row[col]
            if v is not None and str(v) not in ("", "nan", "None"):
                try:
                    result[col] = float(v)
                except Exception:
                    pass
        return result

    # use_max: pick the row with the largest absolute value for each period
    result = {}
    for col in date_cols:
        best = None
        for _, row in rows.iterrows():
            v = row[col]
            if v is not None and str(v) not in ("", "nan", "None"):
                try:
                    fv = float(v)
                    if best is None or abs(fv) > abs(best):
                        best = fv
                except Exception:
                    pass
        if best is not None:
            result[col] = best
    return result


def safe_div(a, b):
    if a is None or b is None or b == 0:
        return None
    return round(a / b * 100, 2)


def compute_metrics(ticker: str, filing_date: str, sector: str) -> dict:
    """
    Load parquets for ticker, compute all metrics, return flat dict.
    Returns {"ticker": ..., "error": ...} on failure.
    """
    base = {"ticker": ticker, "sector": sector, "filing_date": filing_date}

    # Load parquets
    try:
        is_df = pd.read_parquet(os.path.join(XBRL_DIR, f"{ticker}_{filing_date}_income_statement.parquet"))
        bs_df = pd.read_parquet(os.path.join(XBRL_DIR, f"{ticker}_{filing_date}_balance_sheet.parquet"))
        cf_df = pd.read_parquet(os.path.join(XBRL_DIR, f"{ticker}_{filing_date}_cash_flow.parquet"))
    except Exception as e:
        return {**base, "error": f"parquet load failed: {e}"}

    # Get date columns (periods)
    meta_cols = {"standard_concept", "concept", "label", "unit"}
    date_cols = [c for c in is_df.columns if c not in meta_cols]
    if not date_cols:
        return {**base, "error": "no period columns found"}

    # Most recent period = first date column
    p0 = date_cols[0]
    p1 = date_cols[1] if len(date_cols) > 1 else None
    p2 = date_cols[2] if len(date_cols) > 2 else None

    def val(df, concept, period=p0, use_max=False):
        row = get_row_values(df, concept, use_max=use_max)
        if not row:
            return None
        # Use supplied period if it exists in this df's columns.
        # Balance sheet / cash flow parquets may use slightly different
        # period labels than the income statement — fall back to first key.
        if period in row:
            return row[period]
        # Fallback: return value from the first (most recent) period in this df
        return next(iter(row.values()), None)

    # Pull raw values
    # Revenue: use max across all concepts (mirrors _get_max_vector in data_layer)
    # Critical for REITs where RentalAndLeasingRevenue > Revenues concept
    _rev_candidates = [
        val(is_df, "Revenue"), val(is_df, "Revenues"),
        val(is_df, "RentalAndLeasingRevenue"),
        val(is_df, "RegulatedAndUnregulatedOperatingRevenue"),
        val(is_df, "ElectricUtilityRevenue"),
        val(is_df, "RegulatedUtilityRevenue"),
        val(is_df, "UtilitiesRevenue"),
    ]
    rev   = max((v for v in _rev_candidates if v is not None), default=None)
    gp    = val(is_df, "GrossProfit")
    op    = val(is_df, "OperatingIncomeLoss")
    ni    = (val(is_df, "NetIncome") or val(is_df, "NetIncomeToCommonShareholders") or
          val(is_df, "IncomeLossContinuingOperations") or val(is_df, "ProfitLoss"))
    # EBITDA = OperatingIncome + D&A  (mirrors data_layer.py computation)
    # D&A sourced from CF statement add-backs, same concept chain as pipeline
    _da = (val(cf_df, "DepreciationAmortizationCF") or
           val(cf_df, "DepreciationExpense") or
           val(cf_df, "DepreciationAndAmortization") or
           val(cf_df, "DepreciationDepletionAndAmortization") or
           val(cf_df, "DepreciationAmortizationAndAccretionNet") or
           val(cf_df, "OtherDepreciationAndAmortization") or
           val(cf_df, "Depreciation"))
    ebitda = (op + abs(_da)) if (op is not None and _da is not None) else None
    ta    = val(bs_df, "Assets", use_max=True)
    eq    = (val(bs_df, "AllEquityBalance", use_max=True) or
          val(bs_df, "CommonEquity", use_max=True) or
          val(bs_df, "AllEquityBalanceIncludingMinorityInterest", use_max=True))
    td    = ((val(bs_df, "LongTermDebt") or 0) +
          (val(bs_df, "ShortTermDebt") or 0) +
          (val(bs_df, "CurrentPortionOfLongTermDebt") or 0)) or None
    ca    = val(bs_df, "CurrentAssetsTotal")
    cl    = val(bs_df, "CurrentLiabilitiesTotal")
    cash  = (val(bs_df, "CashAndMarketableSecurities") or val(bs_df, "CashAndCashEquivalents"))
    ocf   = val(cf_df, "NetCashFromOperatingActivities", use_max=True)
    capex = (val(cf_df, "CapitalExpenses") or val(cf_df, "PropertyPlantAndEquipmentAdditions"))
    dil   = (val(is_df, "SharesFullyDilutedAverage") or val(is_df, "SharesAverage") or
          val(is_df, "CommonSharesOutstanding") or val(is_df, "SharesOutstanding"))

    # Prior year for CAGR
    rev_p1 = max((v for v in [
        val(is_df, "Revenue", p1), val(is_df, "Revenues", p1),
        val(is_df, "RentalAndLeasingRevenue", p1),
        val(is_df, "RegulatedAndUnregulatedOperatingRevenue", p1),
    ] if v is not None), default=None) if p1 else None
    rev_p2 = max((v for v in [
        val(is_df, "Revenue", p2), val(is_df, "Revenues", p2),
        val(is_df, "RentalAndLeasingRevenue", p2),
        val(is_df, "RegulatedAndUnregulatedOperatingRevenue", p2),
    ] if v is not None), default=None) if p2 else None

    # Compute margins
    gm   = safe_div(gp, rev)
    opm  = safe_div(op, rev)
    nm   = safe_div(ni, rev)
    ebitda_m = safe_div(ebitda, rev)

    # FCF = OCF - |capex|  (mirrors data_layer.py line 1167)
    # capex stored as negative outflow by most filers; abs() normalises both signs.
    # OCF must use use_max=True (see above) to avoid subtotal rows.
    fcf = None
    if ocf is not None and capex is not None and ocf > 0:
        fcf = ocf - abs(capex)
    fcf_m = safe_div(fcf, rev)

    # Returns
    roe = safe_div(ni, eq)
    roa = safe_div(ni, ta)

    # Leverage
    de  = round(td / eq, 2)  if td and eq and eq != 0 else None
    cr  = round(ca / cl, 2)  if ca and cl and cl != 0 else None
    net_debt = (td - cash) if td and cash else td
    nd_ebitda = round(net_debt / ebitda, 2) if (ebitda and ebitda != 0 and net_debt is not None) else None

    # Revenue CAGR (2yr)
    cagr_2yr = None
    if rev and rev_p2 and rev_p2 != 0:
        cagr_2yr = round(((rev / rev_p2) ** 0.5 - 1) * 100, 2)

    # Period labels (clean)
    def clean_period(p):
        return p[:10] if p else None

    # ── Data quality flag ────────────────────────────────────────────────
    # Classifies each ticker's data quality for downstream research use.
    # clean       : all key metrics resolved, delta within expected range
    # gaap_adj    : GAAP figures may diverge from consensus adjusted (Healthcare,
    #               Materials with one-time charges) — flag for analyst note
    # investigate : specific known issue requires diagnose_tags.py
    # reit_rev    : REIT revenue denominator needs verification
    # no_data     : revenue or net income could not be resolved
    dq_flag = "clean"
    dq_note = ""

    if rev is None or ni is None:
        dq_flag = "no_data"
        dq_note = "Revenue or net income unresolvable"
    elif rev < 5e7:  # < $50M revenue — margins meaningless
        dq_flag = "investigate"
        dq_note = f"Revenue ${rev/1e6:.1f}M too small for margin analysis; likely concept mismatch"
    elif sector in ("Healthcare",) and rev and rev > 1e9:
        # Large healthcare companies commonly have IPR&D / litigation charges
        # that create GAAP vs adjusted divergence > 5pp
        dq_flag = "gaap_adj"
        dq_note = "GAAP as-filed; one-time charges (IPR&D/litigation) may diverge from consensus by 5-15pp"
    elif sector == "Real Estate":
        if rev and rev < 2e9:
            dq_flag = "reit_rev"
            dq_note = "REIT revenue <$2B; verify denominator includes all lease income"
        else:
            dq_flag = "gaap_adj"
            dq_note = "REIT — NM includes gains on property sales; FFO basis preferred"
    elif sector == "Materials" and ni is not None and rev is not None:
        nm_val = ni / rev * 100 if rev else 0
        if abs(nm_val) > 20:
            dq_flag = "investigate"
            dq_note = "High margin variance — likely impairment charges; run diagnose_tags.py"

    return {
        **base,
        "period_fy":       clean_period(p0),
        "period_fy1":      clean_period(p1),
        "period_fy2":      clean_period(p2),
        "error":           None,

        # Income statement ($B)
        "revenue_b":       round(rev / 1e9, 3)   if rev   else None,
        "gross_profit_b":  round(gp  / 1e9, 3)   if gp    else None,
        "op_income_b":     round(op  / 1e9, 3)   if op    else None,
        "net_income_b":    round(ni  / 1e9, 3)   if ni    else None,
        "ebitda_b":        round(ebitda / 1e9, 3) if ebitda else None,

        # Margins (%)
        "gross_margin_pct":    gm,
        "operating_margin_pct":opm,
        "net_margin_pct":      nm if nm is None or abs(nm) <= 500 else None,
        "ebitda_margin_pct":   ebitda_m,
        "fcf_margin_pct":      fcf_m,

        # Returns (%)
        "roe_pct": roe if roe is None or abs(roe) <= 500 else None,
        "roa_pct": roa,

        # Leverage
        "de_ratio":       de,
        "current_ratio":  cr,
        "net_debt_b":     round(net_debt / 1e9, 3) if net_debt else None,
        "net_debt_ebitda":nd_ebitda,

        # Growth
        "revenue_cagr_2yr_pct": cagr_2yr,

        # Prior year margins for comparison
        "gross_margin_fy1":    safe_div(val(is_df, "GrossProfit", p1),
                           (val(is_df, "Revenue", p1) or val(is_df, "Revenues", p1))) if p1 else None,
        "operating_margin_fy1":safe_div(val(is_df, "OperatingIncomeLoss", p1),
                           (val(is_df, "Revenue", p1) or val(is_df, "Revenues", p1))) if p1 else None,
        "net_margin_fy1":      safe_div(
                           (val(is_df, "NetIncome", p1) or val(is_df, "NetIncomeToCommonShareholders", p1) or
                            val(is_df, "IncomeLossContinuingOperations", p1) or val(is_df, "ProfitLoss", p1)),
                           (val(is_df, "Revenue", p1) or val(is_df, "Revenues", p1))) if p1 else None,
        "gross_margin_fy2":    safe_div(val(is_df, "GrossProfit", p2),
                           (val(is_df, "Revenue", p2) or val(is_df, "Revenues", p2))) if p2 else None,
        "operating_margin_fy2":safe_div(val(is_df, "OperatingIncomeLoss", p2),
                           (val(is_df, "Revenue", p2) or val(is_df, "Revenues", p2))) if p2 else None,
        "net_margin_fy2":      safe_div(
                           (val(is_df, "NetIncome", p2) or val(is_df, "NetIncomeToCommonShareholders", p2) or
                            val(is_df, "IncomeLossContinuingOperations", p2) or val(is_df, "ProfitLoss", p2)),
                           (val(is_df, "Revenue", p2) or val(is_df, "Revenues", p2))) if p2 else None,

        # Per-share
        "diluted_shares_m": round(dil / 1e6, 1) if dil else None,
        "data_quality":    dq_flag,
        "dq_note":         dq_note,
    }


# ── Load ticker list from cache DB ───────────────────────────────────────────

def load_cached_tickers() -> list[dict]:
    """Return list of {ticker, filing_date} dicts from filing_cache.db."""
    try:
        with sqlite3.connect(CACHE_DB) as con:
            con.row_factory = sqlite3.Row
            rows = con.execute("""
                SELECT ticker, filing_date, form_type
                FROM filings
                ORDER BY ticker
            """).fetchall()
        return [dict(r) for r in rows]
    except Exception as e:
        print(f"ERROR: could not read {CACHE_DB}: {e}")
        sys.exit(1)


def get_sector(ticker: str) -> str:
    """Look up sector from sp500_tickers.py if available."""
    try:
        from sp500_tickers import SP500_DEDUPED
        lookup = {t: s for t, s in SP500_DEDUPED}
        return lookup.get(ticker.upper(), "Unknown")
    except Exception:
        return "Unknown"


# ── Excel writer ─────────────────────────────────────────────────────────────

def write_excel(records: list[dict], failures: list[dict], out_path: str):
    ok = [r for r in records if not r.get("error")]
    df = pd.DataFrame(ok)

    # Column order for summary sheet
    summary_cols = [
        "ticker", "sector", "filing_date", "period_fy",
        "revenue_b",
        "gross_margin_pct", "gross_margin_fy1", "gross_margin_fy2",
        "operating_margin_pct", "operating_margin_fy1", "operating_margin_fy2",
        "net_margin_pct", "net_margin_fy1", "net_margin_fy2",
        "ebitda_margin_pct", "fcf_margin_pct",
        "roe_pct", "roa_pct",
        "de_ratio", "current_ratio",
        "net_debt_b", "net_debt_ebitda",
        "revenue_cagr_2yr_pct",
        "diluted_shares_m",
        "data_quality",
        "dq_note",
    ]
    summary_cols = [c for c in summary_cols if c in df.columns]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:

        # ── Sheet 1: Summary ──
        df[summary_cols].sort_values(["sector", "ticker"]).to_excel(
            writer, sheet_name="Summary", index=False
        )

        # ── Sheet 2: Gross Margin trend ──
        gm_cols = ["ticker", "sector", "revenue_b",
                   "gross_margin_fy2", "gross_margin_fy1", "gross_margin_pct"]
        gm_cols = [c for c in gm_cols if c in df.columns]
        df[gm_cols].sort_values("gross_margin_pct", ascending=False).to_excel(
            writer, sheet_name="Gross Margin", index=False
        )

        # ── Sheet 3: Operating Margin trend ──
        om_cols = ["ticker", "sector", "revenue_b",
                   "operating_margin_fy2", "operating_margin_fy1", "operating_margin_pct"]
        om_cols = [c for c in om_cols if c in df.columns]
        df[om_cols].sort_values("operating_margin_pct", ascending=False).to_excel(
            writer, sheet_name="Operating Margin", index=False
        )

        # ── Sheet 4: Net Margin trend ──
        nm_cols = ["ticker", "sector", "revenue_b",
                   "net_margin_fy2", "net_margin_fy1", "net_margin_pct"]
        nm_cols = [c for c in nm_cols if c in df.columns]
        df[nm_cols].sort_values("net_margin_pct", ascending=False).to_excel(
            writer, sheet_name="Net Margin", index=False
        )

        # ── Sheet 5: Returns ──
        ret_cols = ["ticker", "sector", "roe_pct", "roa_pct",
                    "net_margin_pct", "revenue_cagr_2yr_pct"]
        ret_cols = [c for c in ret_cols if c in df.columns]
        df[ret_cols].sort_values("roe_pct", ascending=False).to_excel(
            writer, sheet_name="Returns", index=False
        )

        # ── Sheet 6: Leverage ──
        lev_cols = ["ticker", "sector", "de_ratio", "current_ratio",
                    "net_debt_b", "net_debt_ebitda", "revenue_b"]
        lev_cols = [c for c in lev_cols if c in df.columns]
        df[lev_cols].sort_values("de_ratio", ascending=False).to_excel(
            writer, sheet_name="Leverage", index=False
        )

        # ── Sheet 7: EBITDA & FCF ──
        cf_cols = ["ticker", "sector", "revenue_b",
                   "ebitda_margin_pct", "fcf_margin_pct", "net_margin_pct"]
        cf_cols = [c for c in cf_cols if c in df.columns]
        df[cf_cols].sort_values("ebitda_margin_pct", ascending=False).to_excel(
            writer, sheet_name="EBITDA & FCF", index=False
        )

        # ── Sheet 8: Revenue Growth ──
        gr_cols = ["ticker", "sector", "revenue_b",
                   "revenue_cagr_2yr_pct", "gross_margin_pct", "operating_margin_pct"]
        gr_cols = [c for c in gr_cols if c in df.columns]
        df[gr_cols].sort_values("revenue_cagr_2yr_pct", ascending=False).to_excel(
            writer, sheet_name="Revenue Growth", index=False
        )

        # ── Sheet 9: Failures ──
        if failures:
            pd.DataFrame(failures).to_excel(
                writer, sheet_name="Failures", index=False
            )

        # ── Apply basic formatting ──
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="2C3E50")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        pct_cols_names = {
            "gross_margin_pct","gross_margin_fy1","gross_margin_fy2",
            "operating_margin_pct","operating_margin_fy1","operating_margin_fy2",
            "net_margin_pct","net_margin_fy1","net_margin_fy2",
            "ebitda_margin_pct","fcf_margin_pct",
            "roe_pct","roa_pct","revenue_cagr_2yr_pct",
        }

        for sheet in writer.book.worksheets:
            # Header row
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Column widths + number formats
            for col_idx, cell in enumerate(sheet[1], 1):
                col_letter = get_column_letter(col_idx)
                col_name   = str(cell.value or "")
                sheet.column_dimensions[col_letter].width = max(12, len(col_name) + 4)
                # Apply % format to margin columns
                if col_name in pct_cols_names:
                    for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for c in row:
                            if c.value is not None:
                                c.number_format = '0.00"%"'

            # Freeze header row
            sheet.freeze_panes = "A2"

    print(f"\nExcel saved → {out_path}")
    print(f"  {len(ok)} tickers across 8 sheets")
    if failures:
        print(f"  {len(failures)} failures logged in 'Failures' sheet")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--out",    type=str, default=DEFAULT_OUT)
    p.add_argument("--ticker", nargs="+", type=str, help="subset of tickers")
    args = p.parse_args()

    cached = load_cached_tickers()
    print(f"Found {len(cached)} tickers in filing_cache.db")

    if args.ticker:
        subset = {t.upper() for t in args.ticker}
        cached = [r for r in cached if r["ticker"] in subset]
        print(f"Filtered to {len(cached)} tickers")

    records  = []
    failures = []
    n = len(cached)

    for i, row in enumerate(cached, 1):
        ticker      = row["ticker"]
        filing_date = row["filing_date"]
        sector      = get_sector(ticker)

        sys.stdout.write(f"\r  Processing {i}/{n}: {ticker:<8} {filing_date}")
        sys.stdout.flush()

        result = compute_metrics(ticker, filing_date, sector)
        if result.get("error"):
            failures.append(result)
        else:
            records.append(result)

    print(f"\n\nComputed: {len(records)} OK  |  {len(failures)} failed")

    # Print quick summary
    ok_df = pd.DataFrame([r for r in records if not r.get("error")])
    if not ok_df.empty and "net_margin_pct" in ok_df.columns:
        nm = ok_df["net_margin_pct"].dropna()
        print(f"\nNet margin distribution across {len(nm)} tickers:")
        print(f"  Median:  {nm.median():.1f}%")
        print(f"  Mean:    {nm.mean():.1f}%")
        print(f"  Min:     {nm.min():.1f}%  ({ok_df.loc[nm.idxmin(),'ticker']})")
        print(f"  Max:     {nm.max():.1f}%  ({ok_df.loc[nm.idxmax(),'ticker']})")
        print(f"  <0%:     {(nm < 0).sum()} tickers")

    if failures:
        print(f"\nFailures:")
        for f in failures:
            print(f"  {f['ticker']:<8}  {f.get('error','')}")

    write_excel(records, failures, args.out)


if __name__ == "__main__":
    main()